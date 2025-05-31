from flask import Flask, render_template, request, redirect, session, url_for, flash, send_file
import os
import requests
from werkzeug.utils import secure_filename
from datetime import datetime, date
from openpyxl import load_workbook
import io
from PyPDF2 import PdfReader, PdfWriter
from supabase import create_client, Client

# --- Configuración inicial ---
app = Flask(__name__)
app.secret_key = 'clave_super_segura'
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'pdf', 'docx', 'doc', 'xls', 'xlsx'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
PDF_BASE = 'FORMULARIO TIPO NEUROLOGIA INFANTIL EDITABLE.pdf'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# --- Supabase ---
SUPABASE_URL = "https://xxxxxxxx.supabase.co"
SUPABASE_KEY = "xxxxx"
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# --- Usuarios simulados ---
USUARIOS = {
    'admin': {'password': 'admin123', 'establecimientos': []},
    'doctora1': {'password': '1234', 'establecimientos': ['Escuela A', 'Liceo B']},
    'doctora2': {'password': 'abcd', 'establecimientos': []}
}

# --- Eventos simulados ---
EVENTOS = [
    {'fecha': '20/05/2025', 'horario': '09:00 - 10:30', 'establecimiento': 'Escuela A', 'obs': 'Evaluación inicial'},
    {'fecha': '21/05/2025', 'horario': '11:00 - 12:30', 'establecimiento': 'Liceo B', 'obs': 'Entrega de informes'}
]

# --- Funciones auxiliares ---
def permitido(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def calculate_age(birth_date):
    today = date.today()
    years = today.year - birth_date.year
    months = today.month - birth_date.month
    if months < 0:
        years -= 1
        months += 12
    return f"{years} años con {months} meses"

def guess_gender(name):
    name = name.lower()
    if name.endswith("a"):
        return "F"
    return "M"

# --- Rutas ---
@app.route('/')
def index():
    return render_template('login.html')

@app.route('/login', methods=['POST'])
def login():
    usuario = request.form['username']
    clave = request.form['password']
    if usuario in USUARIOS and USUARIOS[usuario]['password'] == clave:
        session['usuario'] = usuario
        return redirect(url_for('dashboard'))
    flash('Usuario o contraseña incorrecta')
    return redirect(url_for('index'))

@app.route('/dashboard')
def dashboard():
    if 'usuario' not in session:
        return redirect(url_for('index'))
    usuario = session['usuario']
    establecimientos = USUARIOS[usuario]['establecimientos']
    return render_template('dashboard.html', usuario=usuario, establecimientos=establecimientos, eventos=EVENTOS)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

@app.route('/admin/agregar', methods=['POST'])
def admin_agregar():
    nombre = request.form['nombre']
    fecha = request.form['fecha']
    horario = request.form['horario']
    obs = request.form['obs']
    doctora = request.form['doctora']
    archivo = request.files['formulario']

    if archivo and permitido(archivo.filename):
        filename = secure_filename(f'{nombre}.pdf')
        archivo.save(os.path.join('static/formularios', filename))
        if doctora in USUARIOS and nombre not in USUARIOS[doctora]['establecimientos']:
            USUARIOS[doctora]['establecimientos'].append(nombre)
        EVENTOS.append({'fecha': fecha, 'horario': horario, 'establecimiento': nombre, 'obs': obs})
    return redirect(url_for('dashboard'))

@app.route('/subir/<establecimiento>', methods=['POST'])
def subir(establecimiento):
    if 'usuario' not in session:
        return redirect(url_for('index'))
    archivos = request.files.getlist('archivo')
    if not archivos or archivos[0].filename == '':
        return 'No se seleccionó ningún archivo.', 400
    mensajes = []
    for archivo in archivos:
        if permitido(archivo.filename):
            filename = secure_filename(f"{session['usuario']}_{establecimiento}_{archivo.filename}")
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            archivo.save(filepath)
            mensajes.append(f'✔ {archivo.filename}')
        else:
            mensajes.append(f'✖ {archivo.filename} (no permitido)')
    enviar_correo_sendgrid(
        asunto=f'Nuevos formularios desde {establecimiento}',
        cuerpo=f'Doctora: {session["usuario"]}\nEstablecimiento: {establecimiento}\nSe subieron {len(mensajes)} archivo(s).'
    )
    return "Archivos procesados:<br>" + "<br>".join(mensajes)

@app.route('/evaluados/<establecimiento>', methods=['POST'])
def evaluados(establecimiento):
    if 'usuario' not in session:
        return redirect(url_for('index'))
    cantidad = request.form.get('alumnos')
    usuario = session['usuario']
    enviar_correo_sendgrid(
        asunto=f'Alumnos evaluados - {establecimiento}',
        cuerpo=f'Doctora: {usuario}\nEstablecimiento: {establecimiento}\nCantidad evaluada: {cantidad}'
    )
    return f'Datos enviados correctamente: {cantidad} alumnos evaluados.'

@app.route('/relleno_formularios', methods=['GET', 'POST'])
def relleno_formularios():
    if request.method == 'POST':
        establecimiento = request.form['establecimiento']
        file = request.files['excel']
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        wb = load_workbook(filepath)
        ws = wb.active

        estudiantes = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            nombre, rut, fecha_nac_str, nacionalidad = row
            fecha_nac = datetime.strptime(str(fecha_nac_str), "%d-%m-%Y").date()
            edad = calculate_age(fecha_nac)
            sexo = guess_gender(nombre.split()[0])

            estudiante = {
                'nombre': nombre,
                'rut': rut,
                'fecha_nacimiento': fecha_nac.strftime("%d-%m-%Y"),
                'edad': edad,
                'nacionalidad': nacionalidad,
                'sexo': sexo
            }
            estudiantes.append(estudiante)

        session['estudiantes'] = estudiantes
        return render_template('formulario_relleno.html', estudiantes=estudiantes)

    return render_template('subir_excel.html')

@app.route('/generar_pdf', methods=['POST'])
def generar_pdf():
    datos = request.form
    estudiante_id = int(datos['index'])
    estudiante = session['estudiantes'][estudiante_id]

    diagnostico = datos['diagnostico']
    estado_general = datos['estado_general']
    derivaciones = datos['derivaciones']
    fecha_reevaluacion = datos['fecha_reevaluacion']
    fecha_evaluacion = date.today().strftime("%d-%m-%Y")

    reader = PdfReader(PDF_BASE)
    writer = PdfWriter()
    writer.append(reader)

    data = {
        'nombre': estudiante['nombre'],
        'rut': estudiante['rut'],
        'fecha_nacimiento': estudiante['fecha_nacimiento'],
        'edad': estudiante['edad'],
        'nacionalidad': estudiante['nacionalidad'],
        'diagnostico_1': diagnostico,
        'diagnostico_2': diagnostico,
        'estado_general': estado_general,
        'derivaciones': derivaciones,
        'fecha_evaluacion': fecha_evaluacion,
        'fecha_reevaluacion': fecha_reevaluacion,
        'sexo_f': "Yes" if estudiante['sexo'] == 'F' else "Off",
        'sexo_m': "Yes" if estudiante['sexo'] == 'M' else "Off"
    }

    writer.update_page_form_field_values(writer.pages[0], data)

    pdf_output = io.BytesIO()
    writer.write(pdf_output)
    pdf_output.seek(0)

    return send_file(pdf_output, as_attachment=True, download_name=f"{estudiante['nombre']}_formulario.pdf")

# --- Envío de correos con SendGrid ---
SENDGRID_API_KEY = os.getenv("SENDGRID_API_KEY")
SENDGRID_FROM = 'noreply@cardiohome.cl'
SENDGRID_TO = 'jmiraandal@gmail.com'

def enviar_correo_sendgrid(asunto, cuerpo):
    if not SENDGRID_API_KEY:
        print("Falta SENDGRID_API_KEY en variables de entorno")
        return
    data = {
        "personalizations": [{"to": [{"email": SENDGRID_TO}]}],
        "from": {"email": SENDGRID_FROM},
        "subject": asunto,
        "content": [{"type": "text/plain", "value": cuerpo}]
    }
    try:
        response = requests.post(
            "https://api.sendgrid.com/v3/mail/send",
            headers={
                "Authorization": f"Bearer {SENDGRID_API_KEY}",
                "Content-Type": "application/json"
            },
            json=data
        )
        print(f"Correo enviado, status: {response.status_code}")
    except Exception as e:
        print(f"Error al enviar correo con SendGrid: {e}")

# --- Ejecutar la app ---
if __name__ == '__main__':
    app.run(debug=True)
